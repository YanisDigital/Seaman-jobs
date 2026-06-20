"""Запись вакансий в .xlsx с кликабельными ссылками."""
from __future__ import annotations

import io
import logging
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import COLUMNS, Vacancy

log = logging.getLogger("scraper.excel")

# Ширина колонок по ключу поля (символы).
_WIDTHS = {
    "source": 13, "position": 22, "vessel_type": 20, "fleet": 12,
    "salary": 16, "salary_min": 9, "salary_max": 9, "currency": 8,
    "join_date": 16, "contract_duration": 12, "company": 24,
    "posted": 16, "extra": 18, "description": 60, "vacancy_id": 12,
    "url": 38, "scraped_at": 18,
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_LINK_FONT = Font(color="0563C1", underline="single")


def build_workbook(vacancies: List[Vacancy]) -> Workbook:
    """Собрать книгу Excel: шапка, строки, кликабельные ссылки, автофильтр."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacancies"

    keys = [k for k, _ in COLUMNS]
    headers = [h for _, h in COLUMNS]

    # Заголовок.
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Строки данных.
    for row_idx, vac in enumerate(vacancies, start=2):
        data = asdict(vac)
        for col, key in enumerate(keys, start=1):
            value = data.get(key)
            cell = ws.cell(row=row_idx, column=col)
            if key == "url" and value:
                cell.value = value
                cell.hyperlink = value          # кликабельная ссылка
                cell.font = _LINK_FONT
            else:
                cell.value = value
            if key == "description":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Ширина колонок.
    for col, key in enumerate(keys, start=1):
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS.get(key, 16)

    # Закрепить шапку + автофильтр.
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(keys))
    last_row = len(vacancies) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    return wb


def write_xlsx(vacancies: List[Vacancy], output_dir: str) -> Path:
    """Сохранить вакансии в файл .xlsx (для CLI/GUI). Вернуть путь."""
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = out_dir / f"vacancies_{stamp}.xlsx"
    build_workbook(vacancies).save(path)
    log.info("Saved %d vacancies -> %s", len(vacancies), path)
    return path


def workbook_bytes(vacancies: List[Vacancy]) -> bytes:
    """Книга Excel как байты (для скачивания в вебе, без записи на диск)."""
    buffer = io.BytesIO()
    build_workbook(vacancies).save(buffer)
    return buffer.getvalue()
